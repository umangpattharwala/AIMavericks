"""Document loader for NexaCore RAG documents."""
import os
from pathlib import Path
from typing import Optional

from langchain_core.documents import Document
from langchain_community.document_loaders import (
    Docx2txtLoader,
    CSVLoader,
)
from langchain_community.document_loaders.excel import UnstructuredExcelLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter

from app.config import get_settings


# Mapping of document files to their metadata categories
DOCUMENT_METADATA = {
    "1. Employee_Handbook.docx": {
        "category": "general",
        "scope": "all",
        "description": "General employee handbook covering company culture, code of conduct, workplace policies",
    },
    "2. NEXACORE_Leave_Attendance_Policy.docx": {
        "category": "leave_attendance",
        "scope": "all",
        "description": "Leave types, attendance requirements, holiday calendar",
    },
    "3. Payroll_Policy_Document.docx": {
        "category": "compensation",
        "scope": "all",
        "description": "Salary structure, payroll cycles, tax deductions, bonuses",
    },
    "4. Corporate_Reimbursement_Policy.docx": {
        "category": "reimbursement",
        "scope": "all",
        "description": "Travel, medical, relocation, and other reimbursement policies",
    },
    "5. Employee_Onboarding_Guide.docx": {
        "category": "onboarding",
        "scope": "all",
        "description": "New hire orientation, first-day procedures, probation",
    },
    "6. HR_Employee_QA_200_Questions.docx": {
        "category": "faq",
        "scope": "all",
        "description": "Historical employee questions and HR answers - common queries",
    },
    "7. Employee_Benefits_Policy.docx": {
        "category": "benefits",
        "scope": "all",
        "description": "Healthcare, insurance, wellness programs, stock options, retirement plans",
    },
    "8. HR_Compliance_Security_Policy.docx": {
        "category": "compliance",
        "scope": "hr_only",
        "description": "Data protection, compliance frameworks, security protocols",
    },
    "9.Employee_Lifecycle_Policy.docx": {
        "category": "lifecycle",
        "scope": "all",
        "description": "Promotions, transfers, separations, performance reviews",
    },
    "10. hr_support_tickets.xlsx": {
        "category": "support_tickets",
        "scope": "hr_only",
        "description": "Historical HR support tickets - past issues and resolutions",
    },
}


def load_all_documents(directory: Optional[str] = None) -> list[Document]:
    """Load all RAG documents with metadata enrichment."""
    settings = get_settings()
    doc_dir = Path(directory or settings.rag_documents_dir)
    all_docs = []

    for filename, meta in DOCUMENT_METADATA.items():
        filepath = doc_dir / filename
        if not filepath.exists():
            print(f"Warning: {filename} not found, skipping...")
            continue

        try:
            docs = _load_single_file(filepath)
            # Enrich with metadata
            for doc in docs:
                doc.metadata.update({
                    "source_file": filename,
                    "category": meta["category"],
                    "access_scope": meta["scope"],
                    "description": meta["description"],
                })
            all_docs.extend(docs)
        except Exception as e:
            print(f"Error loading {filename}: {e}")

    return all_docs


def _load_single_file(filepath: Path) -> list[Document]:
    """Load a single file based on its extension."""
    ext = filepath.suffix.lower()

    if ext == ".docx":
        loader = Docx2txtLoader(str(filepath))
    elif ext == ".xlsx":
        return _load_excel_file(filepath)
    elif ext == ".csv":
        loader = CSVLoader(str(filepath))
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    return loader.load()


def _load_excel_file(filepath: Path) -> list[Document]:
    """Load Excel file using openpyxl directly."""
    import openpyxl
    wb = openpyxl.load_workbook(str(filepath), read_only=True)
    docs = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        for row in rows[1:]:
            row_data = {headers[i]: str(cell) if cell is not None else "" for i, cell in enumerate(row)}
            content = " | ".join(f"{k}: {v}" for k, v in row_data.items() if v)
            if content.strip():
                docs.append(Document(
                    page_content=content,
                    metadata={"source": str(filepath), "sheet": sheet_name},
                ))
    wb.close()
    return docs


def chunk_documents(
    documents: list[Document],
    chunk_size: int = 1000,
    chunk_overlap: int = 200,
) -> list[Document]:
    """Split documents into chunks for embedding."""
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        separators=["\n\n", "\n", ". ", " ", ""],
        length_function=len,
    )
    return splitter.split_documents(documents)
